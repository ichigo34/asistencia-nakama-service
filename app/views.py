from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .models import Empleado, TipoAsistencia, RegistroAsistencia, DispositivoEmpleado
from .services import AsistenciaService, ReporteService
from .qr_service import QRService
from .utils import obtener_fecha_hora_actual
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import json


def es_staff(user):
    """
    Verifica si el usuario es staff (administrador).
    
    Args:
        user: Usuario a verificar
        
    Returns:
        bool: True si es staff y está autenticado
    """
    return user.is_authenticated and user.is_staff

@user_passes_test(es_staff)
def pagina_descarga_excel(request):
    """
    Página para descargar reportes de Excel.
    Solo accesible para usuarios staff.
    """
    return render(request, 'pagina_descarga_excel.html')


@user_passes_test(es_staff)
def exportar_resumen_excel(request):
    """
    Exporta un resumen diario de asistencia en formato Excel.
    Delega el procesamiento de datos al ReporteService.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Diario"

    # Encabezados
    encabezados = [
        "Empleado", "Fecha", "Tiempo de Almuerzo",
        "Horas por Comisión", "Horas por Permiso (Otros)",
        "Horas Trabajadas Totales"
    ]
    ws.append(encabezados)

    # Obtener datos usando el servicio
    datos_diarios = ReporteService.obtener_datos_resumen()

    for (id_empleado, fecha), data in datos_diarios.items():
        empleado = data["empleado"]
        horas = ReporteService.calcular_horas_empleado(data)

        ws.append([
            empleado.nombre_completo,
            fecha.strftime("%Y-%m-%d"),
            horas['almuerzo'],
            horas['comision'],
            horas['permiso'],
            horas['trabajadas']
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Aplicar estilo al encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Crear tabla
    tabla = Table(
        displayName="ResumenAsistencia",
        ref=f"A1:F{ws.max_row}"
    )
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    tabla.tableStyleInfo = style
    ws.add_table(tabla)

    # Enviar archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=resumen_asistencia.xlsx'
    wb.save(response)
    return response

@user_passes_test(es_staff)
def exportar_asistencia_excel(request):
    """
    Exporta todos los registros de asistencia en formato Excel.
    Incluye información detallada de cada registro.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Encabezados del excel
    encabezados = ["Empleado", "Tipo de Asistencia", "Fecha", "Hora", "Descripción", "ID Dispositivo"]
    ws.append(encabezados)

    # Registros usando el modelo
    registros = RegistroAsistencia.objects.select_related('empleado', 'tipo') \
        .order_by('-fecha_registro', '-hora_registro')
    
    for reg in registros:
        fila = [
            reg.empleado.nombre_completo,
            reg.tipo.nombre_asistencia,
            reg.fecha_registro.strftime('%Y-%m-%d'),
            reg.hora_registro.strftime('%H:%M:%S'),
            reg.descripcion or '',
            reg.fingerprint or '',
        ]
        ws.append(fila)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Tabla
    tabla = Table(
        displayName="RegistroAsistencia",
        ref=f"A1:F{ws.max_row + 1}"
    )
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    tabla.tableStyleInfo = style
    ws.add_table(tabla)

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=registro_asistencia.xlsx'
    wb.save(response)
    return response



def pagina_principal(request):
    """
    Página principal con opciones de acceso.
    """
    return render(request, 'pagina_principal.html')


def escanear_qr(request):
    """
    Página para escanear código QR.
    """
    return render(request, 'escanear_qr.html')


def identificar_dispositivo(request):
    """
    Página de QR general: identifica por fingerprint. Si ya está vinculado, redirige directo al formulario.
    Si no, muestra selector de empleado para vincular el dispositivo.
    """
    empleados = Empleado.objects.order_by('apellidos', 'nombres')
    return render(request, 'identificar.html', { 'empleados': empleados })


def registrar_asistencia_qr(request, codigo_qr):
    """
    Vista para registrar asistencia usando código QR.
    Detecta automáticamente al empleado.
    """
    empleado = Empleado.buscar_por_codigo_qr(codigo_qr)
    
    if not empleado:
        messages.error(request, 'Código QR no válido o empleado no encontrado.')
        return render(request, 'error_qr.html')
    
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        # Validar datos requeridos
        if not tipo_id:
            messages.error(request, 'Debe seleccionar un tipo de asistencia.')
            return render(request, 'formulario_qr.html', {
                'empleado': empleado,
                'tipos_evento': tipos_evento
            })

        # Usar el servicio para crear el registro
        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado.id_empleado, tipo_id, descripcion, fingerprint
        )

        if success:
            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        else:
            messages.error(request, message)

    return render(request, 'formulario_qr.html', {
        'empleado': empleado,
        'tipos_evento': tipos_evento
    })


def registrar_asistencia_auto(request, empleado_id):
    """
    Registro usando identificación automática por fingerprint (QR general).
    Primera vez: se vincula en identificar_dispositivo.
    """
    empleado = get_object_or_404(Empleado, id_empleado=empleado_id)
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        if not tipo_id:
            messages.error(request, 'Debe seleccionar un tipo de asistencia.')
            return render(request, 'formulario_qr.html', {
                'empleado': empleado,
                'tipos_evento': tipos_evento
            })

        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado.id_empleado, tipo_id, descripcion, fingerprint
        )

        if success:
            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        else:
            messages.error(request, message)

    return render(request, 'formulario_qr.html', {
        'empleado': empleado,
        'tipos_evento': tipos_evento
    })


@csrf_exempt
@require_http_methods(["POST"])
def api_buscar_empleado_qr(request):
    """
    API para buscar empleado por código QR.
    """
    try:
        data = json.loads(request.body)
        codigo_qr = data.get('codigo_qr')
        
        if not codigo_qr:
            return JsonResponse({
                'success': False,
                'error': 'Código QR requerido'
            })
        
        resultado = QRService.buscar_empleado_por_qr(codigo_qr)
        return JsonResponse(resultado)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error del servidor: {str(e)}'
        })


@csrf_exempt
@require_http_methods(["POST"])
def api_identificar_por_fingerprint(request):
    """
    Identifica empleado por fingerprint del dispositivo.
    """
    try:
        data = json.loads(request.body)
        fingerprint = data.get('fingerprint')
        if not fingerprint:
            return JsonResponse({'success': False, 'error': 'Fingerprint requerido'})
        empleado = DispositivoEmpleado.obtener_empleado_por_fingerprint(fingerprint)
        if empleado:
            return JsonResponse({
                'success': True,
                'empleado': {
                    'id': empleado.id_empleado,
                    'nombres': empleado.nombres,
                    'apellidos': empleado.apellidos,
                    'nombre_completo': empleado.nombre_completo,
                    'dni': empleado.dni,
                }
            })
        else:
            return JsonResponse({'success': False, 'error': 'Dispositivo no vinculado a un empleado'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'})


@csrf_exempt
@require_http_methods(["POST"])
def api_vincular_fingerprint(request):
    """
    Vincula el fingerprint al empleado seleccionado (primera vez).
    """
    try:
        data = json.loads(request.body)
        empleado_id = data.get('empleado_id')
        fingerprint = data.get('fingerprint')
        if not empleado_id or not fingerprint:
            return JsonResponse({'success': False, 'error': 'Empleado y fingerprint requeridos'})
        empleado = Empleado.objects.get(id_empleado=empleado_id)
        # Evitar duplicados de fingerprint
        DispositivoEmpleado.objects.update_or_create(
            fingerprint=fingerprint,
            defaults={'empleado': empleado}
        )
        return JsonResponse({'success': True, 'empleado_id': empleado.id_empleado})
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'})


def registrar_asistencia(request):
    """
    Vista tradicional para registrar la asistencia de un empleado.
    Mantenida para compatibilidad.
    """
    empleados = Empleado.objects.all()
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        empleado_id = request.POST.get('empleado')
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        # Validar datos requeridos
        if not empleado_id or not tipo_id:
            messages.error(request, 'Debe seleccionar un empleado y tipo de asistencia.')
            return render(request, 'formulario.html', {
                'empleados': empleados,
                'tipos_evento': tipos_evento
            })

        # Usar el servicio para crear el registro
        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado_id, tipo_id, descripcion, fingerprint
        )

        if success:
            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        else:
            messages.error(request, message)

    return render(request, 'formulario.html', {
        'empleados': empleados,
        'tipos_evento': tipos_evento
    })
